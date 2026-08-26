from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CABECALHO = ["Apartamento", "Hóspede", "Documento", "Check-in", "Check-out", "Observação"]

COR_HOTEL = "0D6EFD"
COR_CABECALHO_BG = "F8F9FA"
COR_TIPO_BG = "EBF3FF"


def gerar_excel_roomlist(grupo, roomlist_agrupada: list[tuple[str, list]]) -> bytes:
    """
    RN-G033: exportação da roomlist completa em Excel (.xlsx), agrupada
    por tipo de apartamento na ordem do cadastro (`ordem`) — o agrupamento
    vem pronto de GrupoRoomlistService.agrupar_por_tipo. Linha ainda vazia
    (hospede_nome nulo) aparece como "A definir".
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Roomlist"

    linha = 1
    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(CABECALHO))
    cel = ws.cell(row=linha, column=1, value=f"{grupo.nome} — {grupo.codigo}")
    cel.font = Font(bold=True, size=13, color=COR_HOTEL)
    linha += 1

    ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(CABECALHO))
    periodo = f'{grupo.data_entrada.strftime("%d/%m/%Y")} a {grupo.data_saida.strftime("%d/%m/%Y")}'
    cel = ws.cell(row=linha, column=1, value=periodo)
    cel.font = Font(size=10, color="6C757D")
    linha += 2

    for nome_tipo, itens in roomlist_agrupada:
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=len(CABECALHO))
        cel = ws.cell(row=linha, column=1, value=f"{nome_tipo} ({len(itens)})")
        cel.font = Font(bold=True, size=11, color=COR_HOTEL)
        cel.fill = PatternFill("solid", fgColor=COR_TIPO_BG)
        linha += 1

        for col, titulo in enumerate(CABECALHO, start=1):
            cel = ws.cell(row=linha, column=col, value=titulo)
            cel.font = Font(bold=True, size=9)
            cel.fill = PatternFill("solid", fgColor=COR_CABECALHO_BG)
        linha += 1

        for item in itens:
            hospede = item.hospede_nome if item.hospede_nome else "A definir"
            if item.cortesia:
                hospede += " (cortesia)"
            valores = [
                item.apartamento or "—",
                hospede,
                item.documento or "—",
                item.check_in.strftime("%d/%m/%Y") if item.check_in else "—",
                item.check_out.strftime("%d/%m/%Y") if item.check_out else "—",
                item.observacao or "—",
            ]
            for col, valor in enumerate(valores, start=1):
                ws.cell(row=linha, column=col, value=valor).font = Font(size=9)
            linha += 1

        linha += 1

    larguras = [14, 28, 18, 12, 12, 30]
    for col, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
